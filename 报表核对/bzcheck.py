import pandas as pd
import openpyxl
import os
from shutil import copyfile
import numpy as np
# excel 数据样式设置类
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class Report:
    def __init__(self):
        self.empty_template_path = '模版/空模版.xlsx'
        self.formula_template_path = '模版/模版公式.xlsx'
        self.result_path = '结果.xlsx'
        self.c_file_name = os.listdir('中国机')[0]
        self.a_file_name = os.listdir('美国机')[0]
        self.dataDict = {
            'c_D27': '',
            'c_D73': '',
            'c_D74': '',
            'a_D27': '',
            'a_D73': '',
            'a_D74': '',
        }
        self.key_word = ['FA0', 'FA1', 'FA3 Category', 'FA3', 'FA0 Category', 'FA1 Category', 'Mfg Site']
        self.function_category = np.asarray(
            ['UI Triage Skipped', 'System', 'Power', 'Face ID', 'Display', 'CND', 'Camera', 'Acoustic',
             'Cellular', 'Safety', 'Activation', 'SIM', 'Sensor', 'WiFi', 'Arc', 'BT', 'Battery', 'Other', 'User'])
        self.exterior_category = ['Cosmetic', 'Mech']
        self.exception_words = ['Pending', 'Under','nan']
        self.note_height = 25
        self.mfg_site = ['FXZZ', 'FXGL', 'LXKS']

        self.load_excel()

    def load_excel(self):
        # copyfile(self.empty_template_path, self.result_path)
        self.result_template = openpyxl.load_workbook(self.empty_template_path)
        self.dataDict['c_D27'] = pd.read_excel('中国机/' + self.c_file_name, sheet_name='D27', header=0)  # 读取Excel文档
        self.dataDict['c_D73'] = pd.read_excel('中国机/' + self.c_file_name, sheet_name='D73', header=0)
        self.dataDict['c_D74'] = pd.read_excel('中国机/' + self.c_file_name, sheet_name='D74', header=0)
        self.dataDict['a_D27'] = pd.read_excel('美国机/' + self.a_file_name, sheet_name='D27', header=0)  # 读取Excel文档
        self.dataDict['a_D73'] = pd.read_excel('美国机/' + self.a_file_name, sheet_name='D73', header=0)
        self.dataDict['a_D74'] = pd.read_excel('美国机/' + self.a_file_name, sheet_name='D74', header=0)
        for key in self.dataDict.keys():
            self.Pretreatment_excel(self.dataDict[key])
        # self.c_D37= pd.read_excel('中国机/' + self.c_file_name, sheet_name=self.sheets_name['D73'], header=None)
        # self.c_D74 = pd.read_excel('中国机/' + self.c_file_name, sheet_name=self.sheets_name['D74'], header=None)
        # self.a_D27 = pd.read_excel('美国机/' + self.a_file_name, sheet_name=self.sheets_name['D27'],header=None)
        # self.a_D37 = pd.read_excel('美国机/' + self.a_file_name, sheet_name=self.sheets_name['D73'], header=None)
        # self.a_D74 = pd.read_excel('美国机/' + self.a_file_name, sheet_name=self.sheets_name['D74'], header=None)
        # self.dataDict['c_D27']=np.asarray(list(c_workbook['D27'].iter_rows(values_only=True)))
        # self.dataDict['c_D73']= list(c_workbook['D73'].iter_rows(values_only=True))
        # self.dataDict['c_D74']= list(c_workbook['D74'].iter_rows(values_only=True))
        # a_workbook = load_workbook(filename='美国机/' + self.a_file_name)
        # self.dataDict['a_D27']= list(a_workbook['D27'].iter_rows(values_only=True))
        # self.dataDict['a_D73']= list(a_workbook['D73'].iter_rows(values_only=True))
        # self.dataDict['a_D74']= list(a_workbook['D74'].iter_rows(values_only=True))
        self.formula_template = list(
            load_workbook(filename=self.formula_template_path)['sheel1'].iter_rows(max_col=23, max_row=81,
                                                                                   values_only=True))
        # self.formula_template = pd.read_excel(self.formula_template_path, sheet_name='sheel1', header=None,).iloc[:, :23]

    def Pretreatment_excel(self, table):
        for key in self.key_word:
            table[key] = table[key].map(lambda x: str(x).strip(), na_action=None)
            table['type'] = table.apply(lambda x: self.getType(x['FA1 Category']), axis=1)

    def iterat_formula(self):
        for i, row in enumerate(self.formula_template):
            for j, value in enumerate(row):
                if not value:
                    continue
                if type(value) == str and value.startswith('//'):
                    # value.replace('/','').trim()
                    formulas = value.replace('/', '').strip().split(';')
                    if formulas[0] == 'FA0Sort':
                        value = self.FA0Sort(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)
                    elif formulas[0] == 'FA1Sort':
                        value, notes, height = self.FA1Sort(*formulas)
                        print(value, formulas, i, j)
                        comment = Comment(notes, "vivian")
                        comment.height = height
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value).comment = comment
                    elif formulas[0] == 'FA2Sort':
                        value = self.FA2Sort(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)

                    elif formulas[0] == 'FA3Sort':
                        value, notes, height = self.FA3Sort(*formulas)
                        print(value, formulas, i, j)
                        comment = Comment(notes, "vivian")
                        comment.height = height
                        comment.width = 300
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value).comment = comment

                    elif formulas[0] == 'MLBSort':
                        value = self.MLBSort(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)
                    elif formulas[0] == 'qtyCount':
                        value = self.qtyCount(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)
                    elif formulas[0] == 'siteCount':
                        value = self.siteCount(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)
                    elif formulas[0] == 'rootCause':
                        value = self.rootCause(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)

                    elif formulas[0] == 'qtyTotal':
                        value = self.qtyTotal(*formulas)
                        print(value, formulas, i, j)
                        self.result_template['sheel1'].cell(row=i + 1, column=j + 1, value=value)

        self.result_template.save(self.result_path)
        print('程序执行成功')



    def getType(self, category):
        if category in self.exterior_category:
            return 'e'
        elif category in self.function_category:
            return 'f'
        else:
            return 'no'

    def FA0Sort(self, *parms):
        table = self.dataDict[parms[1]]
        # table['FA0 Category'] = table['FA0 Category'].map(str.strip, na_action=None)
        type = parms[2]
        number = 0
        if type == 'e':
            number = (table['FA0 Category'].isin(self.exterior_category)).sum()
            return number
        else:
            number = (table['FA0 Category'].isin(self.function_category)).sum()
            return number

    def FA1Sort(self, *parms):
        table = self.dataDict[parms[1]]
        type = parms[2]
        number = 0
        notes = ''
        if type == 'e':
            number = ((table['type'] == 'e') & (table['FA1'] != None) & (
                    table['FA1'].str.contains('|'.join(self.exception_words)) == False)).sum()
            notes_dict = table[(table['type'] == 'e')]['FA1 Category'].value_counts()
            for key, value in notes_dict.items():
                notes += '{}*{}\r\n'.format(key, value)
        else:
            number = ((table['type'] == 'f') & (table['FA1'] != None) & (
                    table['FA1'].str.contains('|'.join(self.exception_words)) == False)).sum()

            notes_dict = table[(table['type'] == 'f')]['FA1 Category'].value_counts()
            for key, value in notes_dict.items():
                notes += '{}*{}\r\n'.format(key, value)
        return number, notes, len(notes_dict) * self.note_height + 10

    def FA2Sort(self, *parms):
        table = self.dataDict[parms[1]]
        type = parms[2]
        number = 0
        notes = ''
        if type == 'e':
            number = ((table['type'] == 'e') & (table['FA2'] != None) & (
                    table['FA2'].str.contains('|'.join(self.exception_words)) == False)).sum()
        else:
            number = ((table['type'] == 'f') & (table['FA2'] != None) & (
                    table['FA2'].str.contains('|'.join(self.exception_words)) == False)).sum()
        return number

    def FA3Sort(self, *parms):
        table = self.dataDict[parms[1]]
        type = parms[2]
        number = 0
        notes = ''
        if type == 'e':
            number = ((table['type'] == 'e') & (table['FA3'] != None) & (
                    table['FA3'].str.contains('|'.join(self.exception_words)) == False)).sum()
            fa1_notes_dict = table[((table['type'] == 'e') & (table['FA2'] == 'MLB') & (
                table['FA3'].str.contains('|'.join(self.exception_words))))]['FA1'].value_counts()
            fa3_notes_dict = table[((table['type'] == 'e') & (table['FA2'] == 'MLB') & (
                    table['FA3'].str.contains('|'.join(self.exception_words)) == False))]['FA3'].value_counts()
            notes += 'FA3:\r\n'
            for key, value in fa3_notes_dict.items():
                notes += '{}*{}\r\n'.format(key, value)
            notes += 'Under FA:\r\n'
            for key, value in fa1_notes_dict.items():
                notes += '{}*{}\r\n'.format(key, value)
        else:
            number = ((table['type'] == 'f') & (table['FA3'] != None) & (
                    table['FA3'].str.contains('|'.join(self.exception_words)) == False)).sum()
            fa1_notes_dict = table[((table['type'] == 'f') & (table['FA2'] == 'MLB') & (
                table['FA3'].str.contains('|'.join(self.exception_words))))]['FA1'].value_counts()
            fa3_notes_dict = table[((table['type'] == 'f') & (table['FA2'] == 'MLB') & (
                        table['FA3'].str.contains('|'.join(self.exception_words)) == False))]['FA3'].value_counts()
            notes += 'FA3:\r\n'
            for key, value in fa3_notes_dict.items():
                notes += '{}*{}\r\n'.format(key, value)
            notes += 'Under FA:\r\n'
            for key, value in fa1_notes_dict.items():
                notes += '{}*{}\r\n'.format(key, value)
        return number, notes, (len(fa1_notes_dict) + len(fa3_notes_dict)) * self.note_height + 40

    def MLBSort(self, *parms):
        table = pd.concat([self.dataDict[key] for key in (parms[1].split(','))])
        # print(table)
        type = parms[2]
        number = 0
        notes = ''
        fa3_dict = ''
        if type == 'e':
            fa3_dict = table[((table['type'] == 'e') & (table['FA2'] == 'MLB'))]['FA3'].value_counts()
        elif type == 'f':
            fa3_dict = table[((table['type'] == 'f') & (table['FA2'] == 'MLB'))]['FA3'].value_counts()
        else:
            fa3_dict = table[(table['FA2'] == 'MLB')]['FA3'].value_counts()
        for key, value in fa3_dict.items():
            notes += '{}*{}\r\n'.format(key, value)
        return notes.strip()

    def qtyCount(self, *parms):
        table = self.dataDict[parms[1]]
        type = parms[2]
        number = 0
        if type == 'm':
            number = ((table['FA3'] != None) & (table['FA2'] == 'MLB') & (
                        table['FA3'].str.contains('|'.join(self.exception_words)) == False)).sum()
        elif type == 'u':
            number = ((table['FA2'] == 'MLB') & (
                        (table['FA3'] == None) | (table['FA3'].str.contains('|'.join(self.exception_words))))).sum()

        return number

    def siteCount(self, *parms):
        table = self.dataDict[parms[1]]
        type = parms[2]
        size_str = ''
        if type == 'm':
            for site in self.mfg_site:
                number = ((table['FA3'] != None) & (table['FA2'] == 'MLB') & (
                        table['FA3'].str.contains('|'.join(self.exception_words)) == False) & (
                                  table['Mfg Site'] == site)).sum()
                if number > 0:
                    size_str += '{}*{}\r\n'.format(site, number)
        elif type == 'u':
            for site in self.mfg_site:
                number = ((table['FA2'] == 'MLB') & (table['Mfg Site'] == site) & (
                        (table['FA3'] == None) | (table['FA3'].str.contains('|'.join(self.exception_words))))).sum()
                if number > 0:
                    size_str += '{}*{}\r\n'.format(site, number)

        return size_str.strip()

    def rootCause(self, *parms):
        table = self.dataDict[parms[1]]
        type = parms[2]
        cause_str = ''
        if type == 'm':
            for site in self.mfg_site:
                site_dict = table[((table['FA3'] != None) & (table['FA2'] == 'MLB') & (
                        table['FA3'].str.contains('|'.join(self.exception_words)) == False) & (
                                               table['Mfg Site'] == site))]['FA3'].value_counts()
                site_fa = ''
                sum=0
                for key, value in site_dict.items():
                    sum+=value
                    site_fa += '{}*{}、'.format(key, value)
                if sum>0:
                    cause_str += '{}: {}\r\n'.format(site, site_fa)
        elif type == 'u':
            for site in self.mfg_site:
                site_dict = table[((table['FA2'] == 'MLB') & (table['Mfg Site'] == site) & (
                            (table['FA3'] == None) | (table['FA3'].str.contains('|'.join(self.exception_words)))))][
                    'FA1'].value_counts()
                site_fa = ''
                sum = 0
                for key, value in site_dict.items():
                    sum += value
                    site_fa += '{}*{}、'.format(key, value)
                if sum > 0:
                    cause_str += '{}: {}\r\n'.format(site, site_fa)
        return cause_str.strip()

    def qtyTotal(self,*parms):
        table = pd.concat([self.dataDict[key] for key in (parms[1].split(','))])
        # print(table)
        type = parms[2]
        total_str=''
        if type=='fxzz':
            number1 = ((table['FA3'] != None) & (table['FA2'] == 'MLB') & (
                    table['FA3'].str.contains('|'.join(self.exception_words)) == False) & (
                              table['Mfg Site'] == self.mfg_site[0])).sum()
            number2= ((table['FA2'] == 'MLB') & (table['Mfg Site'] == self.mfg_site[0]) & (
                    (table['FA3'] == None) | (table['FA3'].str.contains('|'.join(self.exception_words))))).sum()
            total_str+='Material*{}\r\nUnder FA*{}\r\nProcess*0'.format(number1,number2)
        elif type == 'fxgl':
            number1 = ((table['FA3'] != None) & (table['FA2'] == 'MLB') & (
                    table['FA3'].str.contains('|'.join(self.exception_words)) == False) & (
                               table['Mfg Site'] == self.mfg_site[1])).sum()
            number2 = ((table['FA2'] == 'MLB') & (table['Mfg Site'] == self.mfg_site[1]) & (
                    (table['FA3'] == None) | (table['FA3'].str.contains('|'.join(self.exception_words))))).sum()
            total_str += 'Material*{}\r\nUnder FA*{}\r\nProcess*0'.format(number1, number2)
        elif type=='lxks':
            number1 = ((table['FA3'] != None) & (table['FA2'] == 'MLB') & (
                    table['FA3'].str.contains('|'.join(self.exception_words)) == False) & (
                               table['Mfg Site'] == self.mfg_site[2])).sum()
            number2 = ((table['FA2'] == 'MLB') & (table['Mfg Site'] == self.mfg_site[2]) & (
                    (table['FA3'] == None) | (table['FA3'].str.contains('|'.join(self.exception_words))))).sum()
            total_str += 'Material*{}\r\nUnder FA*{}\r\nProcess*0'.format(number1, number2)

        return total_str

if __name__ == "__main__":
    r = Report()
    r.iterat_formula()
